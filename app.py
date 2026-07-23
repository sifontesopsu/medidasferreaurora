import io
import json
import re
import threading
import time
import unicodedata
import uuid
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
import requests
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from async_sync import DurableSyncQueue, post_json


st.set_page_config(page_title="Control Medidas ML", page_icon="📦", layout="wide")

APPS_SCRIPT_URL = "https://script.google.com/macros/s/AKfycbyOWOLy003wtbEtcW59QYUYkPboPj1gBhKKZi5s-xV23KcgKgzAn-FHIqTVxC2bRzAj8w/exec"
SYNC_DATA_DIR = Path(".data").expanduser()
SYNC_DB_PATH = str(SYNC_DATA_DIR / "sync_queue.sqlite3")
SYNC_SPOOL_DIR = str(SYNC_DATA_DIR / "sync_spool")
ESTADOS_CIERRE = [
    "listo_para_ejecutivo",
    "en_gestion_ejecutivo",
    "resuelto",
    "rechazado_ml",
    "rechazado_ejecutivo",
]
BANDEJAS_ADMINISTRATIVA = {
    "Pendientes por gestionar": ["validado_supervisor"],
    "Enviados a ejecutiva": ["listo_para_ejecutivo"],
    "En gestión ejecutiva": ["en_gestion_ejecutivo"],
    "Cerrados": ["resuelto", "rechazado_ml", "rechazado_ejecutivo"],
}


# =========================================================
# API
# =========================================================
def current_api_auth() -> Dict[str, str]:
    user = st.session_state.get("auth_user")
    if not isinstance(user, dict):
        return {}
    usuario_id = str(user.get("usuario_id", "") or "").strip()
    return {"usuario_id": usuario_id} if usuario_id else {}


def api_post(payload: Dict[str, Any], timeout: int = 180) -> Dict[str, Any]:
    body = dict(payload)
    body.setdefault("request_id", str(uuid.uuid4()))
    auth = current_api_auth()
    if auth and body.get("action") not in {"login_with_pin"}:
        body["auth"] = auth
    return post_json(
        APPS_SCRIPT_URL,
        body,
        timeout=timeout,
    )


@st.cache_resource
def get_sync_queue(
    endpoint: str,
    db_path: str,
    spool_dir: str,
) -> DurableSyncQueue:
    queue = DurableSyncQueue(
        endpoint=endpoint,
        db_path=db_path,
        spool_dir=spool_dir,
        poll_seconds=0.8,
        request_timeout=180,
        max_attempts=8,
    )
    queue.cleanup(done_older_than_days=7)
    return queue.start()


SYNC_QUEUE = get_sync_queue(
    APPS_SCRIPT_URL,
    SYNC_DB_PATH,
    SYNC_SPOOL_DIR,
)


def enqueue_api_action(
    action: str,
    payload: Dict[str, Any],
    *,
    entity_key: str = "",
) -> Dict[str, Any]:
    body = dict(payload)
    auth = current_api_auth()
    if auth:
        body["auth"] = auth
    job_id = SYNC_QUEUE.enqueue(action, body, entity_key=entity_key)
    return {"ok": True, "queued": True, "job_id": job_id}


@st.cache_data(ttl=25, show_spinner=False)
def api_get_dashboard_counts() -> Dict[str, Any]:
    return api_post({"action": "get_dashboard_counts"}, timeout=120)


@st.cache_data(ttl=15, show_spinner=False)
def api_get_admin_queue(
    query: str = "",
    estados: Optional[List[str]] = None,
    operador: str = "",
    limit: int = 1000000,
) -> pd.DataFrame:
    data = api_post(
        {
            "action": "get_admin_queue",
            "query": query,
            "estados": estados or [],
            "operador": operador,
            "limit": limit,
        },
        timeout=120,
    )
    return pd.DataFrame(data.get("items", []))


@st.cache_data(ttl=15, show_spinner=False)
def api_get_admin_queue_grouped_by_sku(
    query: str = "",
    estados: Optional[List[str]] = None,
    operador: str = "",
) -> pd.DataFrame:
    data = api_post(
        {
            "action": "get_admin_queue_grouped_by_sku",
            "query": query,
            "estados": estados or [],
            "operador": operador,
        },
        timeout=120,
    )
    return pd.DataFrame(data.get("items", []))

@st.cache_data(ttl=15, show_spinner=False)
def api_get_admin_snapshot(
    query: str = "",
    estados: Optional[List[str]] = None,
    operador: str = "",
) -> Dict[str, Any]:
    return api_post(
        {
            "action": "get_admin_snapshot",
            "query": query,
            "estados": estados or [],
            "operador": operador,
        },
        timeout=120,
    )


@st.cache_data(ttl=15, show_spinner=False)
def api_get_tasks_by_operator(operador: str) -> pd.DataFrame:
    data = api_post({"action": "get_tasks_by_operator", "operador": operador}, timeout=120)
    return pd.DataFrame(data.get("items", []))


@st.cache_data(ttl=15, show_spinner=False)
def api_get_tasks_by_operator_grouped_by_sku(operador: str) -> pd.DataFrame:
    data = api_post({"action": "get_tasks_by_operator_grouped_by_sku", "operador": operador}, timeout=120)
    return pd.DataFrame(data.get("items", []))


def api_mark_sku_no_stock(sku: str, operador: str, usuario: str, comentario: str = "") -> Dict[str, Any]:
    result = enqueue_api_action(
        "mark_sku_no_stock",
        {
            "sku": sku,
            "operador": operador,
            "usuario": usuario,
            "comentario": comentario,
        },
        entity_key=f"no_stock:{sku}",
    )
    result["affected"] = "en cola"
    return result

@st.cache_data(ttl=15, show_spinner=False)
def api_get_pending_validation(limit: int = 200) -> pd.DataFrame:
    data = api_post({"action": "get_pending_validation", "limit": limit}, timeout=120)
    return pd.DataFrame(data.get("items", []))


@st.cache_data(ttl=15, show_spinner=False)
def api_get_pending_validation_grouped_by_sku(limit: int = 200) -> pd.DataFrame:
    data = api_post({"action": "get_pending_validation_grouped_by_sku", "limit": limit}, timeout=120)
    return pd.DataFrame(data.get("items", []))


@st.cache_data(ttl=15, show_spinner=False)
def api_get_administrative_queue(statuses: List[str], limit: int = 300) -> pd.DataFrame:
    data = api_post(
        {"action": "get_administrative_queue", "statuses": statuses, "limit": limit},
        timeout=120,
    )
    return pd.DataFrame(data.get("items", []))


@st.cache_data(ttl=20, show_spinner=False)
def api_get_case_detail(sku: str, mlc: str) -> Dict[str, Any]:
    return api_post({"action": "get_case_detail", "sku": sku, "mlc": mlc}, timeout=120)


@st.cache_data(ttl=20, show_spinner=False)
def api_get_case_detail_by_sku(sku: str) -> Dict[str, Any]:
    return api_post({"action": "get_case_detail_by_sku", "sku": sku}, timeout=120)


@st.cache_data(ttl=30, show_spinner=False)
def api_get_evidencias(sku: str, mlc: str) -> pd.DataFrame:
    data = api_post({"action": "get_evidencias", "sku": sku, "mlc": mlc}, timeout=120)
    return pd.DataFrame(data.get("data", []))


@st.cache_data(ttl=30, show_spinner=False)
def api_get_evidencias_by_sku(sku: str) -> pd.DataFrame:
    data = api_post({"action": "get_evidencias_by_sku", "sku": sku}, timeout=120)
    return pd.DataFrame(data.get("data", []))


def api_login_with_pin(usuario: str, pin: str) -> Dict[str, Any]:
    return api_post({"action": "login_with_pin", "usuario": usuario, "pin": pin}, timeout=60)

def api_get_user_profile(usuario_id: str) -> Dict[str, Any]:
    return api_post(
        {"action": "get_user_profile", "usuario_id": usuario_id},
        timeout=60,
    )

@st.cache_data(ttl=300, show_spinner=False)
def api_get_active_operators() -> pd.DataFrame:
    data = api_post({"action": "get_active_operators"}, timeout=60)
    return pd.DataFrame(data.get("data", []))


def api_assign_tasks_grouped_by_sku(items: List[Dict[str, str]], operador: str, usuario: str) -> Dict[str, Any]:
    result = enqueue_api_action(
        "assign_tasks_grouped_by_sku",
        {
            "items": items,
            "operador": operador,
            "usuario": usuario,
        },
        entity_key=f"assign:{operador}",
    )
    result["assigned"] = len(items)
    return result

def api_save_measurement_with_photos_by_sku(
    sku: str,
    operador: str,
    alto_real_cm: float,
    ancho_real_cm: float,
    profundidad_real_cm: float,
    peso_real_kg: float,
    observacion_operador: str,
    foto_alto,
    foto_ancho,
    foto_profundidad,
    foto_peso,
) -> Dict[str, Any]:
    payload = {
        "sku": sku,
        "operador": operador,
        "alto_real_cm": alto_real_cm,
        "ancho_real_cm": ancho_real_cm,
        "profundidad_real_cm": profundidad_real_cm,
        "peso_real_kg": peso_real_kg,
        "observacion_operador": observacion_operador,
        "auth": current_api_auth(),
    }
    job_id = SYNC_QUEUE.enqueue_with_uploaded_files(
        "save_measurement_with_photos_by_sku",
        payload,
        [
            ("alto", foto_alto),
            ("ancho", foto_ancho),
            ("profundidad", foto_profundidad),
            ("peso", foto_peso),
        ],
        entity_key=f"measurement_sku:{sku}",
    )
    return {
        "ok": True,
        "queued": True,
        "job_id": job_id,
        "medicion_id": "pendiente",
        "publicaciones_afectadas": "en cola",
    }

def api_validate_measurement_by_sku(sku: str, supervisor: str, aprobar: bool, comentario: str) -> Dict[str, Any]:
    result = enqueue_api_action(
        "validate_measurement_by_sku",
        {
            "sku": sku,
            "supervisor": supervisor,
            "aprobar": aprobar,
            "comentario": comentario,
        },
        entity_key=f"validation_sku:{sku}",
    )
    result["affected"] = "en cola"
    return result

def api_update_status(
    sku: str,
    mlc: str,
    nuevo_estado: str,
    usuario: str,
    comentario: str = "",
    ticket_ejecutivo: str = "",
) -> Dict[str, Any]:
    return enqueue_api_action(
        "update_status",
        {
            "sku": sku,
            "mlc": mlc,
            "nuevo_estado": nuevo_estado,
            "usuario": usuario,
            "comentario": comentario,
            "ticket_ejecutivo": ticket_ejecutivo,
        },
        entity_key=f"admin_status:{sku}:{mlc}",
    )

def normalize_identifier(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if not text or text.lower() in {"nan", "none", "null"}:
        return ""
    if text.endswith(".0"):
        text = text[:-2]
    return text


INVENTORY_SKU_ALIASES = [
    "sku",
    "codigo sku",
    "codigo producto",
    "codigo articulo",
    "cod producto",
    "cod articulo",
    "codigo",
    "referencia",
    "item",
]
INVENTORY_STOCK_ALIASES = [
    "q saldo consolidado",
    "saldo consolidado",
    "q saldo",
    "saldo",
    "stock",
    "stock real",
    "stock disponible",
    "saldo actual",
    "saldo bodega",
    "existencia",
    "existencias",
    "disponible",
    "cantidad disponible",
    "cantidad",
]


def normalize_inventory_header(value: Any) -> str:
    text = unicodedata.normalize("NFD", str(value or ""))
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    text = re.sub(r"[^a-zA-Z0-9]+", " ", text).strip().lower()
    return re.sub(r"\s+", " ", text)


def guess_inventory_column(columns: List[Any], aliases: List[str]) -> Optional[str]:
    normalized = {str(column): normalize_inventory_header(column) for column in columns}
    aliases_normalized = [normalize_inventory_header(alias) for alias in aliases]

    for alias in aliases_normalized:
        for original, header in normalized.items():
            if header == alias:
                return original
    for alias in aliases_normalized:
        for original, header in normalized.items():
            if alias and (header.startswith(alias + " ") or header.endswith(" " + alias)):
                return original
    for alias in aliases_normalized:
        for original, header in normalized.items():
            if alias and alias in header:
                return original
    return None


def parse_inventory_number(value: Any) -> Optional[float]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, bool):
        return float(int(value))
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace("\u00a0", "").replace(" ", "")
    if not text or text.lower() in {"nan", "none", "null", "-"}:
        return None
    text = re.sub(r"[^0-9,\.\-+]", "", text)
    if not text:
        return None

    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        if re.fullmatch(r"[-+]?\d{1,3}(,\d{3})+", text):
            text = text.replace(",", "")
        else:
            text = text.replace(",", ".")
    elif "." in text and re.fullmatch(r"[-+]?\d{1,3}(\.\d{3})+", text):
        text = text.replace(".", "")

    try:
        return float(text)
    except ValueError:
        return None


@st.cache_data(show_spinner=False, max_entries=3)
def list_inventory_sheets(file_bytes: bytes, filename: str) -> List[str]:
    extension = Path(filename).suffix.lower()
    if extension not in {".xlsx", ".xlsm"}:
        return []
    excel_file = pd.ExcelFile(io.BytesIO(file_bytes), engine="openpyxl")
    return list(excel_file.sheet_names)


def detect_inventory_header_row(preview_df: pd.DataFrame) -> int:
    """Encuentra la fila que contiene simultáneamente una columna SKU y una de stock."""
    if preview_df.empty:
        return 0

    for row_index, row in preview_df.iterrows():
        values = [normalize_inventory_header(value) for value in row.tolist()]
        has_sku = guess_inventory_column(values, INVENTORY_SKU_ALIASES) is not None
        has_stock = guess_inventory_column(values, INVENTORY_STOCK_ALIASES) is not None
        if has_sku and has_stock:
            return int(row_index)
    return 0


@st.cache_data(show_spinner=False, max_entries=3)
def read_inventory_report(file_bytes: bytes, filename: str, sheet_name: Optional[str] = None) -> pd.DataFrame:
    extension = Path(filename).suffix.lower()
    if extension in {".xlsx", ".xlsm"}:
        selected_sheet = sheet_name or 0
        preview = pd.read_excel(
            io.BytesIO(file_bytes),
            sheet_name=selected_sheet,
            header=None,
            nrows=30,
            dtype=object,
            engine="openpyxl",
        )
        header_row = detect_inventory_header_row(preview)
        report = pd.read_excel(
            io.BytesIO(file_bytes),
            sheet_name=selected_sheet,
            header=header_row,
            dtype=object,
            engine="openpyxl",
        )
        report.attrs["detected_header_row"] = header_row + 1
        return report

    if extension == ".csv":
        last_error: Optional[Exception] = None
        for encoding in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                preview = pd.read_csv(
                    io.BytesIO(file_bytes),
                    header=None,
                    nrows=30,
                    dtype=object,
                    sep=None,
                    engine="python",
                    encoding=encoding,
                )
                header_row = detect_inventory_header_row(preview)
                report = pd.read_csv(
                    io.BytesIO(file_bytes),
                    header=header_row,
                    dtype=object,
                    sep=None,
                    engine="python",
                    encoding=encoding,
                )
                report.attrs["detected_header_row"] = header_row + 1
                return report
            except Exception as error:
                last_error = error
        raise RuntimeError(f"No se pudo leer el CSV: {last_error}")
    raise RuntimeError("Formato no compatible. Usa XLSX, XLSM o CSV")


def build_temporary_inventory(raw_df: pd.DataFrame, sku_column: str, stock_column: str) -> pd.DataFrame:
    if raw_df.empty:
        raise RuntimeError("El informe no contiene filas")
    if sku_column not in raw_df.columns or stock_column not in raw_df.columns:
        raise RuntimeError("Las columnas seleccionadas no existen en el informe")

    inventory = raw_df[[sku_column, stock_column]].copy()
    inventory["sku_inventario"] = inventory[sku_column].map(normalize_identifier)
    inventory["stock_inventario"] = inventory[stock_column].map(parse_inventory_number)
    inventory = inventory.loc[inventory["sku_inventario"] != ""].copy()
    inventory = inventory.loc[inventory["stock_inventario"].notna()].copy()
    if inventory.empty:
        raise RuntimeError("No se encontraron SKU y valores de stock válidos")

    inventory["stock_inventario"] = pd.to_numeric(inventory["stock_inventario"], errors="coerce")
    inventory = (
        inventory.groupby("sku_inventario", as_index=False, dropna=False)["stock_inventario"]
        .sum(min_count=1)
        .sort_values("sku_inventario")
        .reset_index(drop=True)
    )
    return inventory


@st.cache_resource
def get_temporary_inventory_store() -> Dict[str, Any]:
    return {
        "lock": threading.RLock(),
        "data": pd.DataFrame(columns=["sku_inventario", "stock_inventario"]),
        "filename": "",
        "sheet_name": "",
        "sku_column": "",
        "stock_column": "",
        "loaded_at": "",
    }


def get_temporary_inventory_snapshot() -> Dict[str, Any]:
    store = get_temporary_inventory_store()
    with store["lock"]:
        data = store.get("data")
        return {
            "loaded": isinstance(data, pd.DataFrame) and not data.empty,
            "data": data.copy() if isinstance(data, pd.DataFrame) else pd.DataFrame(),
            "filename": str(store.get("filename", "")),
            "sheet_name": str(store.get("sheet_name", "")),
            "sku_column": str(store.get("sku_column", "")),
            "stock_column": str(store.get("stock_column", "")),
            "loaded_at": str(store.get("loaded_at", "")),
        }


def set_temporary_inventory(
    inventory_df: pd.DataFrame,
    *,
    filename: str,
    sheet_name: str,
    sku_column: str,
    stock_column: str,
) -> None:
    store = get_temporary_inventory_store()
    with store["lock"]:
        store["data"] = inventory_df.copy()
        store["filename"] = str(filename)
        store["sheet_name"] = str(sheet_name or "")
        store["sku_column"] = str(sku_column)
        store["stock_column"] = str(stock_column)
        store["loaded_at"] = time.strftime("%d-%m-%Y %H:%M:%S")


def clear_temporary_inventory() -> None:
    store = get_temporary_inventory_store()
    with store["lock"]:
        store["data"] = pd.DataFrame(columns=["sku_inventario", "stock_inventario"])
        store["filename"] = ""
        store["sheet_name"] = ""
        store["sku_column"] = ""
        store["stock_column"] = ""
        store["loaded_at"] = ""


def format_inventory_stock(value: Any) -> str:
    if value is None or pd.isna(value):
        return "No encontrado"
    number = float(value)
    if number.is_integer():
        return f"{int(number):,}".replace(",", ".")
    return f"{number:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def api_bulk_update_status(items: List[Dict[str, str]], nuevo_estado: str, usuario: str) -> Dict[str, Any]:
    normalized_items: List[Dict[str, str]] = []
    for item in items:
        sku = normalize_identifier(item.get("sku", ""))
        mlc = normalize_identifier(item.get("mlc", ""))
        if sku and mlc:
            normalized_items.append({"sku": sku, "mlc": mlc})
    if not normalized_items:
        raise RuntimeError("No hay items válidos para actualizar")
    result = enqueue_api_action(
        "bulk_update_status",
        {
            "items": normalized_items,
            "nuevo_estado": nuevo_estado,
            "usuario": usuario,
        },
        entity_key="bulk_price_update",
    )
    result["updated"] = len(normalized_items)
    result["errors"] = 0
    return result


def api_bulk_update_administrative_status(
    items: List[Dict[str, str]],
    nuevo_estado: str,
    usuario: str,
    comentario: str,
    ticket_ejecutivo: str = "",
) -> Dict[str, Any]:
    normalized_items = []
    for item in items:
        sku = normalize_identifier(item.get("sku", ""))
        mlc = normalize_identifier(item.get("mlc", ""))
        if sku and mlc:
            normalized_items.append({"sku": sku, "mlc": mlc})
    if not normalized_items:
        raise RuntimeError("No hay items válidos para actualizar")
    result = enqueue_api_action(
        "bulk_update_administrative_status",
        {
            "items": normalized_items,
            "nuevo_estado": nuevo_estado,
            "usuario": usuario,
            "comentario": comentario,
            "ticket_ejecutivo": ticket_ejecutivo,
        },
        entity_key="bulk_admin_status",
    )
    result["queued_items"] = len(normalized_items)
    return result

# =========================================================
# HELPERS
# =========================================================
def clear_caches(include_session_queues: bool = False) -> None:
    api_get_dashboard_counts.clear()
    api_get_admin_queue.clear()
    api_get_admin_queue_grouped_by_sku.clear()
    api_get_admin_snapshot.clear()
    api_get_tasks_by_operator.clear()
    api_get_tasks_by_operator_grouped_by_sku.clear()
    api_get_pending_validation.clear()
    api_get_pending_validation_grouped_by_sku.clear()
    api_get_administrative_queue.clear()
    api_get_case_detail.clear()
    api_get_case_detail_by_sku.clear()
    api_get_evidencias.clear()
    api_get_evidencias_by_sku.clear()
    api_get_active_operators.clear()
    if include_session_queues:
        for key in (
            "admin_queue_pub_df",
            "admin_queue_sku_df",
            "admin_queue_signature",
            "admin_queue_cached_version",
            "supervisor_queue_df",
            "supervisor_queue_df_version",
        ):
            st.session_state.pop(key, None)
        st.session_state["admin_queue_version"] = st.session_state.get("admin_queue_version", 0) + 1
        st.session_state["supervisor_queue_version"] = st.session_state.get("supervisor_queue_version", 0) + 1


def reconcile_completed_syncs() -> None:
    revision = SYNC_QUEUE.revision()
    previous = st.session_state.get("sync_revision_seen")
    if previous is None:
        st.session_state["sync_revision_seen"] = revision
        return
    if int(previous) != int(revision):
        clear_caches(include_session_queues=True)
        st.session_state["sync_revision_seen"] = revision


def remember_pending(namespace: str, entity: str, job_id: str) -> None:
    key = f"pending_sync_{namespace}"
    values = st.session_state.setdefault(key, {})
    values[str(entity)] = str(job_id)


def active_pending_entities(namespace: str) -> set[str]:
    key = f"pending_sync_{namespace}"
    values = st.session_state.get(key, {})
    if not isinstance(values, dict):
        return set()
    active: Dict[str, str] = {}
    for entity, job_id in values.items():
        job = SYNC_QUEUE.get_job(str(job_id))
        if job and job.get("status") in {"pending", "processing"}:
            active[str(entity)] = str(job_id)
    st.session_state[key] = active
    return set(active.keys())


def get_allowed_admin_status_transitions(estado_actual: str) -> List[str]:
    estado_actual = str(estado_actual or "").strip()
    if estado_actual == "validado_supervisor":
        return ["listo_para_ejecutivo", "en_gestion_ejecutivo", "resuelto", "rechazado_ml", "rechazado_ejecutivo"]
    if estado_actual == "listo_para_ejecutivo":
        return ["en_gestion_ejecutivo", "resuelto", "rechazado_ml", "rechazado_ejecutivo"]
    if estado_actual == "en_gestion_ejecutivo":
        return ["resuelto", "rechazado_ml", "rechazado_ejecutivo"]
    if estado_actual in ["resuelto", "rechazado_ml", "rechazado_ejecutivo"]:
        return [estado_actual]
    return []


def validate_admin_status_change(
    estado_actual: str,
    nuevo_estado: str,
    comentario: str,
    ticket_ejecutivo: str = "",
) -> Optional[str]:
    estado_actual = str(estado_actual or "").strip()
    nuevo_estado = str(nuevo_estado or "").strip()
    comentario = str(comentario or "").strip()
    ticket_ejecutivo = str(ticket_ejecutivo or "").strip()

    if not comentario:
        return "El comentario es obligatorio"

    if estado_actual in ["resuelto", "rechazado_ml", "rechazado_ejecutivo"]:
        return "El caso ya está cerrado"

    permitidos = get_allowed_admin_status_transitions(estado_actual)
    if nuevo_estado not in permitidos:
        return f"La transición desde {estado_actual} hacia {nuevo_estado} no está permitida"

    requiere_ticket = nuevo_estado in ["listo_para_ejecutivo", "en_gestion_ejecutivo"]
    if requiere_ticket and not ticket_ejecutivo:
        return "Debes ingresar ticket ejecutivo para este estado"

    return None


def validate_bulk_admin_status_change(
    casos: List[Dict[str, Any]],
    nuevo_estado: str,
    comentario: str,
    ticket_ejecutivo: str = "",
) -> tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    validos: List[Dict[str, Any]] = []
    bloqueados: List[Dict[str, Any]] = []

    for caso in casos:
        error = validate_admin_status_change(
            estado_actual=str(caso.get("estado_actual", "")),
            nuevo_estado=nuevo_estado,
            comentario=comentario,
            ticket_ejecutivo=ticket_ejecutivo,
        )
        if error:
            bloqueados.append({"caso": caso, "motivo": error})
        else:
            validos.append(caso)

    return validos, bloqueados


def refresh_supervisor_queue(limit: int = 300, force: bool = False) -> pd.DataFrame:
    cache_key = "supervisor_queue_df"
    version_key = "supervisor_queue_version"
    current_version = st.session_state.get(version_key, 0)
    cached_df = st.session_state.get(cache_key)
    cached_version = st.session_state.get(f"{cache_key}_version")

    if (not force) and isinstance(cached_df, pd.DataFrame) and cached_version == current_version:
        return cached_df.copy()

    df = safe_df(api_get_pending_validation_grouped_by_sku(limit=limit))
    st.session_state[cache_key] = df.copy()
    st.session_state[f"{cache_key}_version"] = current_version
    return df


def bump_supervisor_queue_version() -> None:
    st.session_state["supervisor_queue_version"] = st.session_state.get("supervisor_queue_version", 0) + 1


def remove_supervisor_sku_from_queue(sku: str) -> None:
    cache_key = "supervisor_queue_df"
    cached_df = st.session_state.get(cache_key)
    if isinstance(cached_df, pd.DataFrame) and not cached_df.empty:
        remaining = cached_df[cached_df["sku"].astype(str) != str(sku)].reset_index(drop=True)
        st.session_state[cache_key] = remaining


def get_admin_filter_signature(query: str, estados: Optional[List[str]], operador: str) -> tuple:
    estados_norm = tuple(sorted(str(x) for x in (estados or [])))
    return (str(query or "").strip(), estados_norm, str(operador or "").strip())


def refresh_admin_queues(
    query: str = "",
    estados: Optional[List[str]] = None,
    operador: str = "",
    force: bool = False,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    signature = get_admin_filter_signature(query, estados, operador)
    version = st.session_state.get("admin_queue_version", 0)
    cached_signature = st.session_state.get("admin_queue_signature")
    cached_version = st.session_state.get("admin_queue_cached_version")
    cached_pub = st.session_state.get("admin_queue_pub_df")
    cached_sku = st.session_state.get("admin_queue_sku_df")

    if (
        not force
        and cached_signature == signature
        and cached_version == version
        and isinstance(cached_pub, pd.DataFrame)
        and isinstance(cached_sku, pd.DataFrame)
    ):
        return cached_pub.copy(), cached_sku.copy()

    snapshot = api_get_admin_snapshot(query=query, estados=estados, operador=operador)
    df_pub = safe_df(pd.DataFrame(snapshot.get("publications", [])))
    df_sku = safe_df(pd.DataFrame(snapshot.get("grouped", [])))

    st.session_state["admin_queue_signature"] = signature
    st.session_state["admin_queue_cached_version"] = version
    st.session_state["admin_queue_pub_df"] = df_pub.copy()
    st.session_state["admin_queue_sku_df"] = df_sku.copy()
    return df_pub, df_sku


def update_admin_queue_after_assignment(selected_skus: List[str], operador_destino: str) -> None:
    selected_skus_str = {str(x) for x in selected_skus}
    if not selected_skus_str:
        return

    operador_filter = str(st.session_state.get("admin_filters_state", {}).get("operador_filter", "") or "").strip()
    keep_rows = (not operador_filter) or (operador_filter.casefold() == str(operador_destino).strip().casefold())

    pub_df = st.session_state.get("admin_queue_pub_df")
    if isinstance(pub_df, pd.DataFrame) and not pub_df.empty and "sku" in pub_df.columns:
        mask_pub = pub_df["sku"].astype(str).isin(selected_skus_str)
        pub_new = pub_df.copy()
        if keep_rows:
            if "operador_asignado" in pub_new.columns:
                pub_new.loc[mask_pub, "operador_asignado"] = str(operador_destino).strip()
        else:
            pub_new = pub_new.loc[~mask_pub].reset_index(drop=True)
        st.session_state["admin_queue_pub_df"] = pub_new

    sku_df = st.session_state.get("admin_queue_sku_df")
    if isinstance(sku_df, pd.DataFrame) and not sku_df.empty and "sku" in sku_df.columns:
        mask_sku = sku_df["sku"].astype(str).isin(selected_skus_str)
        sku_new = sku_df.copy()
        if keep_rows:
            if "operador_asignado" in sku_new.columns:
                sku_new.loc[mask_sku, "operador_asignado"] = str(operador_destino).strip()
        else:
            sku_new = sku_new.loc[~mask_sku].reset_index(drop=True)
        st.session_state["admin_queue_sku_df"] = sku_new


def safe_df(df: pd.DataFrame) -> pd.DataFrame:
    return df if isinstance(df, pd.DataFrame) and not df.empty else pd.DataFrame()


def row_is_no_stock(row: Any) -> bool:
    try:
        sin_stock_value = str(row.get("sin_stock", "")).strip().lower()
        prioridad_value = str(row.get("prioridad", "")).strip().lower().replace("_", " ")
    except Exception:
        return False
    return sin_stock_value in {"true", "1", "si", "sí"} or "sin stock" in prioridad_value


def badge_estado(estado: str) -> str:
    color_map = {
        "pendiente_medicion": "#f59e0b",
        "requiere_nueva_evidencia": "#f97316",
        "medido_pendiente_validacion": "#3b82f6",
        "validado_supervisor": "#10b981",
        "listo_para_ejecutivo": "#6366f1",
        "en_gestion_ejecutivo": "#8b5cf6",
        "resuelto": "#16a34a",
        "rechazado_ml": "#dc2626",
        "rechazado_ejecutivo": "#b91c1c",
    }
    color = color_map.get(str(estado), "#6b7280")
    return (
        f"<span style='background:{color};color:white;padding:4px 8px;"
        f"border-radius:999px;font-size:12px'>{estado}</span>"
    )


def show_kpi_row_from_counts(counts: Dict[str, Any]) -> None:
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.metric("Total", int(counts.get("total", 0)))
    with c2:
        st.metric("Pendiente medición", int(counts.get("pendiente_medicion", 0)))
    with c3:
        st.metric("Pendiente validación", int(counts.get("medido_pendiente_validacion", 0)))
    with c4:
        st.metric("Pendiente gestión administrativa", int(counts.get("validado_supervisor", 0)))


def build_ejecutiva_excel_bytes(df: pd.DataFrame, seller_id: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja 1"

    ws.merge_cells("C1:F1")
    ws["C1"] = "PACKAGING"
    ws["C1"].font = Font(bold=True)
    ws["C1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["C1"].fill = PatternFill(fill_type="solid", fgColor="F4DDC6")

    headers = [
        "seller_id",
        "item_id",
        "height_propuesto (cm)",
        "width_propuesto (cm)",
        "length_propuesto (cm)",
        "weight_propuesto (gr)",
    ]
    header_fills = ["C6E0B4", "C6E0B4", "F4DDC6", "F4DDC6", "F4DDC6", "F4DDC6"]
    thin = Side(style="thin", color="A0A0A0")

    for col_idx, (header, color) in enumerate(zip(headers, header_fills), start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_idx, (_, row) in enumerate(df.iterrows(), start=3):
        ws.cell(row=row_idx, column=1, value=seller_id)
        ws.cell(row=row_idx, column=2, value=str(row.get("mlc", "")))
        ws.cell(row=row_idx, column=3, value=row.get("alto_real_cm", ""))
        ws.cell(row=row_idx, column=4, value=row.get("ancho_real_cm", ""))
        ws.cell(row=row_idx, column=5, value=row.get("profundidad_real_cm", ""))
        peso_kg = pd.to_numeric(pd.Series([row.get("peso_real_kg", "")]), errors="coerce").iloc[0]
        peso_gr = "" if pd.isna(peso_kg) else round(float(peso_kg) * 1000, 2)
        ws.cell(row=row_idx, column=6, value=peso_gr)
        for col_idx in range(1, 7):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    widths = {"A": 14, "B": 18, "C": 20, "D": 20, "E": 22, "F": 22}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A3"

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


def build_comparativas_excel_bytes(df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparativas"

    headers = [
        "sku",
        "mlc",
        "titulo",
        "alto_ml_cm",
        "ancho_ml_cm",
        "profundidad_ml_cm",
        "peso_ml_kg",
        "alto_real_cm",
        "ancho_real_cm",
        "profundidad_real_cm",
        "peso_real_kg",
        "estado_actual",
        "operador_asignado",
        "supervisor",
        "fecha_ultima_medicion",
        "fecha_validacion",
    ]

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="A0A0A0")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    export_df = df.copy()
    for col in headers:
        if col not in export_df.columns:
            export_df[col] = ""

    export_df = export_df[headers]

    for row_idx, (_, row) in enumerate(export_df.iterrows(), start=2):
        for col_idx, value in enumerate(row.tolist(), start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    widths = {
        "A": 18, "B": 16, "C": 55, "D": 12, "E": 12, "F": 16, "G": 12,
        "H": 12, "I": 12, "J": 16, "K": 12, "L": 24, "M": 18, "N": 18,
        "O": 20, "P": 20
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A2"

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


def normalize_case_payload(detail: Dict[str, Any], fallback: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    fallback = fallback or {}
    if not isinstance(detail, dict):
        return fallback.copy()

    for key in ["case", "item", "data", "detail", "row"]:
        value = detail.get(key)
        if isinstance(value, dict) and value:
            merged = fallback.copy()
            merged.update(value)
            return merged

    merged = fallback.copy()
    flat = {k: v for k, v in detail.items() if not isinstance(v, (dict, list))}
    merged.update(flat)
    return merged


def render_case_summary(case: Dict[str, Any]) -> None:
    if not case:
        st.warning("No se pudo cargar el caso")
        return

    c1, c2 = st.columns(2)
    with c1:
        st.markdown(f"**SKU:** {case.get('sku', '')}")
        st.markdown(f"**MLC:** {case.get('mlc', '')}")
        st.markdown(f"**Título:** {case.get('titulo', '')}")
        st.markdown(badge_estado(str(case.get('estado_actual', ''))), unsafe_allow_html=True)
    with c2:
        st.markdown(
            f"**ML:** {case.get('alto_ml_cm', '')} x {case.get('ancho_ml_cm', '')} x {case.get('profundidad_ml_cm', '')} cm | {case.get('peso_ml_kg', '')} kg"
        )
        st.markdown(
            f"**Real:** {case.get('alto_real_cm', '')} x {case.get('ancho_real_cm', '')} x {case.get('profundidad_real_cm', '')} cm | {case.get('peso_real_kg', '')} kg"
        )
        st.markdown(f"**Operador:** {case.get('operador_asignado', '')}")
        st.markdown(f"**Supervisor:** {case.get('supervisor', '')}")

    obs_operador = case.get("observacion_operador", "")
    obs_admin = case.get("observacion_admin", "")
    if obs_operador:
        st.info(f"Observación operador: {obs_operador}")
    if obs_admin:
        st.info(f"Observación admin/supervisor: {obs_admin}")


def build_drive_view_url(row) -> str:
    file_id = str(row.get("drive_file_id", "") or "").strip()
    if file_id:
        return f"https://drive.google.com/uc?export=view&id={file_id}"

    drive_link = str(row.get("drive_link", "") or "").strip()
    if "id=" in drive_link:
        return drive_link

    return drive_link


def render_drive_image(row):
    url = build_drive_view_url(row)
    if not url:
        st.warning("Sin link")
        return

    try:
        headers = {"User-Agent": "Mozilla/5.0"}
        resp = requests.get(url, headers=headers, timeout=30)
        resp.raise_for_status()
        st.image(resp.content, use_container_width=True)
    except Exception:
        st.warning("No se pudo mostrar imagen")


def render_evidencias(evidencias: pd.DataFrame) -> None:
    st.markdown("### Evidencia fotográfica")
    if evidencias.empty:
        st.warning("No hay fotos disponibles")
        return

    evidencias = evidencias.copy()
    if "tipo_foto" not in evidencias.columns:
        st.warning("Las evidencias recibidas no incluyen el campo tipo_foto")
        return
    evidencias["tipo_foto"] = evidencias["tipo_foto"].astype(str).str.lower()
    if "fecha_carga" in evidencias.columns:
        evidencias = evidencias.sort_values("fecha_carga")
    evidencias = evidencias.drop_duplicates(subset=["tipo_foto"], keep="last")

    orden = ["alto", "ancho", "profundidad", "peso"]
    mapa = {str(r["tipo_foto"]).lower(): r for _, r in evidencias.iterrows()}
    cols = st.columns(4)

    for i, tipo in enumerate(orden):
        with cols[i]:
            st.markdown(f"**{tipo.upper()}**")
            row = mapa.get(tipo)
            if row is None:
                st.warning("Falta foto")
                continue
            render_drive_image(row)


def require_login() -> Dict[str, Any]:
    auth_user = st.session_state.get("auth_user")
    if isinstance(auth_user, dict) and auth_user:
        return auth_user

    st.title("Control Medidas ML")
    st.subheader("Ingreso con PIN")
    with st.form("login_form"):
        usuario = st.text_input("Usuario")
        pin = st.text_input("PIN", type="password")
        submitted = st.form_submit_button("Ingresar", use_container_width=True)
    if submitted:
        try:
            auth = api_login_with_pin(usuario, pin)
            user = auth.get("user", {})
            st.session_state["auth_user"] = user
            st.rerun()
        except Exception as e:
            st.error(f"No se pudo iniciar sesión: {e}")
    st.stop()

def toggle_evidencias(case_key: str) -> None:
    current = st.session_state.get(case_key, False)
    st.session_state[case_key] = not current


# =========================================================
# AUTH / SIDEBAR
# =========================================================
if not APPS_SCRIPT_URL:
    st.error("Falta configurar APPS_SCRIPT_URL en app.py")
    st.stop()

reconcile_completed_syncs()
user = require_login()
rol = str(user.get("rol", "")).strip().lower()
usuario_actual = str(user.get("usuario_id", user.get("nombre", "")))
operador_codigo = str(user.get("operador_codigo", "") or usuario_actual)

with st.sidebar:
    st.title("Control Medidas ML")
    st.caption(f"Usuario: {user.get('nombre', usuario_actual)}")
    st.caption(f"Rol: {rol}")
    sync_stats = SYNC_QUEUE.stats()
    st.caption(
        "Sincronización: "
        f"{sync_stats.get('pending', 0)} pendientes | "
        f"{sync_stats.get('processing', 0)} procesando | "
        f"{sync_stats.get('error', 0)} con error"
    )
    inventory_snapshot_sidebar = get_temporary_inventory_snapshot()
    if inventory_snapshot_sidebar.get("loaded"):
        inventory_rows = len(inventory_snapshot_sidebar.get("data", pd.DataFrame()))
        st.success(
            f"Inventario temporal activo: {inventory_rows} SKU\n\n"
            f"Actualizado: {inventory_snapshot_sidebar.get('loaded_at', '')}"
        )
    else:
        st.caption("Inventario temporal: no cargado")

    if rol == "admin":
        with st.expander("Cargar inventario temporal", expanded=False):
            st.caption(
                "Se comparte con todos los operadores mientras la app esté encendida. "
                "No se guarda en Sheets ni en archivos del servidor."
            )
            temporary_inventory_file = st.file_uploader(
                "Informe de inventario",
                type=["xlsx", "xlsm", "csv"],
                key="temporary_inventory_file",
            )
            if temporary_inventory_file is not None:
                try:
                    temporary_file_bytes = temporary_inventory_file.getvalue()
                    temporary_sheet_names = list_inventory_sheets(
                        temporary_file_bytes,
                        temporary_inventory_file.name,
                    )
                    selected_temporary_sheet = ""
                    if temporary_sheet_names:
                        default_sheet_index = 0
                        if "id" in temporary_sheet_names:
                            default_sheet_index = temporary_sheet_names.index("id")
                        selected_temporary_sheet = st.selectbox(
                            "Hoja del inventario",
                            temporary_sheet_names,
                            index=default_sheet_index,
                            key="temporary_inventory_sheet",
                        )
                    temporary_raw_df = read_inventory_report(
                        temporary_file_bytes,
                        temporary_inventory_file.name,
                        selected_temporary_sheet or None,
                    )
                    temporary_raw_df.columns = [str(column) for column in temporary_raw_df.columns]
                    if temporary_raw_df.empty:
                        st.warning("El informe no contiene datos")
                    else:
                        temporary_columns = [str(column) for column in temporary_raw_df.columns]
                        sku_guess = guess_inventory_column(temporary_columns, INVENTORY_SKU_ALIASES)
                        stock_guess = guess_inventory_column(temporary_columns, INVENTORY_STOCK_ALIASES)
                        sku_index = temporary_columns.index(sku_guess) if sku_guess in temporary_columns else 0
                        stock_index = temporary_columns.index(stock_guess) if stock_guess in temporary_columns else 0
                        temporary_sku_column = st.selectbox(
                            "Columna SKU",
                            temporary_columns,
                            index=sku_index,
                            key="temporary_inventory_sku_column",
                        )
                        temporary_stock_column = st.selectbox(
                            "Columna stock",
                            temporary_columns,
                            index=stock_index,
                            key="temporary_inventory_stock_column",
                        )
                        detected_header_row = int(temporary_raw_df.attrs.get("detected_header_row", 1))
                        st.caption(
                            f"Encabezados detectados en la fila {detected_header_row} · "
                            f"Filas de inventario: {len(temporary_raw_df):,}".replace(",", ".")
                        )
                        if st.button("Activar inventario temporal", use_container_width=True):
                            temporary_inventory_df = build_temporary_inventory(
                                temporary_raw_df,
                                temporary_sku_column,
                                temporary_stock_column,
                            )
                            set_temporary_inventory(
                                temporary_inventory_df,
                                filename=temporary_inventory_file.name,
                                sheet_name=selected_temporary_sheet,
                                sku_column=temporary_sku_column,
                                stock_column=temporary_stock_column,
                            )
                            st.success(
                                f"Inventario activado con {len(temporary_inventory_df):,} SKU."
                                .replace(",", ".")
                            )
                            st.rerun()
                except Exception as error:
                    st.error(f"No se pudo cargar el inventario: {error}")

            if inventory_snapshot_sidebar.get("loaded"):
                if st.button("Quitar inventario temporal", use_container_width=True):
                    clear_temporary_inventory()
                    st.success("Inventario temporal eliminado")
                    st.rerun()
    if sync_stats.get("error", 0):
        with st.expander("Ver errores de sincronización"):
            for failure in SYNC_QUEUE.recent_failures(limit=5):
                st.error(
                    f"{failure.get('action')} · {failure.get('entity_key') or failure.get('job_id')}\n\n"
                    f"{failure.get('last_error')}"
                )
            if st.button("Reintentar sincronizaciones fallidas", use_container_width=True):
                retried = SYNC_QUEUE.retry_failed()
                st.success(f"Reintentos activados: {retried}")
                st.rerun()
    if st.button("Recargar desde Sheets", use_container_width=True):
        clear_caches(include_session_queues=True)
        st.rerun()
    if st.button("Cerrar sesión", use_container_width=True):
        st.session_state.clear()
        st.rerun()

if rol == "admin":
    opciones_modulo = ["Administrador", "Operador", "Supervisor", "Administrativa"]
elif rol == "operador":
    opciones_modulo = ["Operador"]
elif rol == "supervisor":
    opciones_modulo = ["Supervisor"]
elif rol == "administrativa":
    opciones_modulo = ["Administrativa"]
else:
    st.error("Tu usuario tiene un rol no reconocido. Solicita al administrador corregirlo en la hoja usuarios.")
    st.stop()

modo = st.sidebar.radio("Módulo", opciones_modulo)


# =========================================================
# ADMINISTRADOR
# =========================================================
if modo == "Administrador":
    st.title("Panel Administrador")

    try:
        counts = api_get_dashboard_counts()
    except Exception as e:
        st.error(f"No se pudieron cargar los indicadores: {e}")
        st.stop()

    show_kpi_row_from_counts(counts)

    st.subheader("Filtros")
    with st.form("admin_filters_form"):
        f1, f2, f3 = st.columns([2, 2, 2])
        with f1:
            texto = st.text_input("Buscar SKU / MLC / título")
        with f2:
            estados_disponibles = counts.get("estados_disponibles", [])
            estados_sel = st.multiselect("Estado", estados_disponibles, default=estados_disponibles)
        with f3:
            operador_filter = st.text_input("Operador asignado")
        filtros_submit = st.form_submit_button("Aplicar filtros", use_container_width=True)

    admin_filter_state = st.session_state.setdefault(
        "admin_filters_state",
        {"texto": "", "estados_sel": counts.get("estados_disponibles", []), "operador_filter": ""},
    )
    if filtros_submit:
        admin_filter_state["texto"] = texto.strip()
        admin_filter_state["estados_sel"] = estados_sel
        admin_filter_state["operador_filter"] = operador_filter.strip()

    try:
        df_filtrado_pub, df_filtrado_sku = refresh_admin_queues(
            query=admin_filter_state["texto"],
            estados=admin_filter_state["estados_sel"],
            operador=admin_filter_state["operador_filter"],
        )
    except Exception as e:
        st.error(f"No se pudo cargar la bandeja administrativa: {e}")
        st.stop()

    df_filtrado_pub = safe_df(df_filtrado_pub)
    df_filtrado_sku = safe_df(df_filtrado_sku)

    if "ventas" not in df_filtrado_sku.columns:
        df_filtrado_sku["ventas"] = df_filtrado_sku.get("ventas_total", "")

    if df_filtrado_sku.empty:
        st.info("No hay SKUs para los filtros seleccionados")
        st.stop()

    st.caption(f"Resultados encontrados: {len(df_filtrado_sku)} SKUs | {len(df_filtrado_pub)} publicaciones")

    st.subheader("Asignación de tareas por SKU")
    try:
        operadores_df = safe_df(api_get_active_operators())
    except Exception as e:
        st.error(f"No se pudo cargar la lista de operadores: {e}")
        st.stop()
    if operadores_df.empty:
        st.warning("No hay usuarios activos con rol operador")
        st.stop()
    operadores_df["label"] = operadores_df.apply(
        lambda row: f"{row.get('nombre', '')} | {row.get('operador_codigo', '')}",
        axis=1,
    )
    operador_label_to_code = dict(zip(operadores_df["label"], operadores_df["operador_codigo"].astype(str)))
    with st.form("admin_assign_form"):
        operador_label = st.selectbox("Asignar a operador", operadores_df["label"].tolist())
        operador_destino = operador_label_to_code.get(operador_label, "")
        cols_view = [c for c in ["sku", "titulo", "ventas", "estado_actual", "operador_asignado", "publicaciones_count"] if c in df_filtrado_sku.columns]
        edited = st.data_editor(
            df_filtrado_sku[cols_view].assign(seleccionar=False),
            use_container_width=True,
            hide_index=True,
            column_config={"seleccionar": st.column_config.CheckboxColumn("Seleccionar", default=False)},
            disabled=cols_view,
            key="admin_editor_asignacion_sku",
        )
        asignar_btn = st.form_submit_button("Asignar SKUs seleccionados", use_container_width=True)

    if asignar_btn:
        if not operador_destino.strip():
            st.warning("Debes indicar el nombre del operador")
        else:
            seleccionados = edited[edited["seleccionar"] == True]  # noqa: E712
            if seleccionados.empty:
                st.warning("No seleccionaste SKUs")
            else:
                items = seleccionados[["sku"]].to_dict(orient="records")
                try:
                    result = api_assign_tasks_grouped_by_sku(items, operador_destino.strip(), usuario_actual)
                    update_admin_queue_after_assignment(
                        [str(x) for x in seleccionados["sku"].astype(str).tolist()],
                        operador_destino.strip(),
                    )
                    st.success(
                        f"Asignación enviada a segundo plano para {result.get('assigned', 0)} SKU(s). "
                        f"Trabajo: {result.get('job_id', '')[:8]}"
                    )
                    st.rerun()
                except Exception as e:
                    st.error(f"No se pudo asignar: {e}")

    st.subheader("Precios resueltos")
    df_resueltos = safe_df(
        api_get_admin_queue(
            query="",
            estados=["resuelto"],
            operador="",
            limit=1000000,
        )
    )

    pending_price_keys = active_pending_entities("price_update")
    if pending_price_keys and not df_resueltos.empty:
        composite = df_resueltos.apply(
            lambda row: f"{normalize_identifier(row.get('sku', ''))}::{normalize_identifier(row.get('mlc', ''))}",
            axis=1,
        )
        df_resueltos = df_resueltos.loc[~composite.isin(pending_price_keys)].reset_index(drop=True)

    if df_resueltos.empty:
        st.info("No hay casos resueltos pendientes de marcar como precio_actualizado")
    else:
        cols_precio = [c for c in ["sku", "mlc", "titulo", "ventas", "estado_actual", "fecha_resolucion"] if c in df_resueltos.columns]
        with st.form("admin_precio_actualizado_form"):
            precio_editor = st.data_editor(
                df_resueltos[cols_precio].assign(seleccionar=False),
                use_container_width=True,
                hide_index=True,
                column_config={"seleccionar": st.column_config.CheckboxColumn("Seleccionar", default=False)},
                disabled=cols_precio,
                key="admin_editor_precio_actualizado",
            )
            marcar_precio_btn = st.form_submit_button("Marcar seleccionados como precio_actualizado", use_container_width=True)

        if marcar_precio_btn:
            seleccionados_precio = precio_editor[precio_editor["seleccionar"] == True]  # noqa: E712
            if seleccionados_precio.empty:
                st.warning("No seleccionaste casos resueltos")
            else:
                items_precio = [
                    {
                        "sku": normalize_identifier(row_precio.get("sku", "")),
                        "mlc": normalize_identifier(row_precio.get("mlc", "")),
                    }
                    for _, row_precio in seleccionados_precio.iterrows()
                ]
                try:
                    result_precio = api_bulk_update_status(items_precio, "precio_actualizado", usuario_actual)
                    for item_precio in items_precio:
                        remember_pending(
                            "price_update",
                            f"{item_precio['sku']}::{item_precio['mlc']}",
                            str(result_precio.get("job_id", "")),
                        )
                    actualizados = int(result_precio.get("updated", 0))
                    st.success(
                        f"Se enviaron {actualizados} cambios de precio a segundo plano. "
                        f"Trabajo: {str(result_precio.get('job_id', ''))[:8]}"
                    )
                    st.rerun()
                except Exception as e:
                    st.error(f"No se pudo marcar precio_actualizado: {e}")

    st.subheader("Reporte comparativo")
    comparativas_bytes = build_comparativas_excel_bytes(df_filtrado_pub if not df_filtrado_pub.empty else df_filtrado_sku)
    st.download_button(
        "Descargar Excel comparativas",
        data=comparativas_bytes,
        file_name="reporte_comparativas_medidas.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=False,
        key="download_comparativas_admin",
    )



# =========================================================
# OPERADOR
# =========================================================
elif modo == "Operador":
    st.title("Módulo Operador PDA")

    if rol == "operador":
        nombre_operador = operador_codigo
        st.text_input("Operador", value=nombre_operador, disabled=True, key="nombre_operador_bloqueado")
    else:
        if st.button("Recargar lista de operadores", key="recargar_operadores_pda", use_container_width=True):
            api_get_active_operators.clear()
            st.rerun()

        operadores_df = safe_df(api_get_active_operators())
        if operadores_df.empty:
            st.warning(
                "No hay operadores activos. Revisa la hoja usuarios: rol=operador, activo=TRUE/1/SI "
                "y operador_codigo con valor."
            )
            st.stop()

        for columna in ["usuario_id", "nombre", "operador_codigo"]:
            if columna not in operadores_df.columns:
                operadores_df[columna] = ""
            operadores_df[columna] = operadores_df[columna].fillna("").astype(str).str.strip()

        operadores_df["operador_codigo"] = operadores_df["operador_codigo"].where(
            operadores_df["operador_codigo"].ne(""), operadores_df["usuario_id"]
        )
        operadores_df["nombre"] = operadores_df["nombre"].where(
            operadores_df["nombre"].ne(""), operadores_df["operador_codigo"]
        )
        operadores_df = (
            operadores_df.loc[operadores_df["operador_codigo"].ne("")]
            .drop_duplicates(subset=["operador_codigo"], keep="first")
            .sort_values(["nombre", "operador_codigo"], kind="stable")
            .reset_index(drop=True)
        )

        if operadores_df.empty:
            st.warning("Los operadores activos no tienen usuario ni código de operador configurado")
            st.stop()

        operadores_df["label"] = operadores_df.apply(
            lambda row: (
                f"{row.get('nombre', '')} | {row.get('operador_codigo', '')}"
                + (f" | {row.get('usuario_id', '')}" if row.get("usuario_id", "") not in {"", row.get("operador_codigo", "")} else "")
            ),
            axis=1,
        )
        operadores_df["_buscar"] = operadores_df.apply(
            lambda row: normalize_inventory_header(
                f"{row.get('nombre', '')} {row.get('operador_codigo', '')} {row.get('usuario_id', '')}"
            ),
            axis=1,
        )

        st.caption(f"Operadores activos encontrados: {len(operadores_df)}")
        busqueda_operador = st.text_input(
            "Buscar operador",
            placeholder="Escribe el nombre, usuario o código del operador",
            key="buscar_operador_admin_pda",
        )
        termino_operador = normalize_inventory_header(busqueda_operador)
        operadores_filtrados = operadores_df
        if termino_operador:
            palabras = [p for p in termino_operador.split(" ") if p]
            mascara = operadores_df["_buscar"].apply(lambda texto: all(p in texto for p in palabras))
            operadores_filtrados = operadores_df.loc[mascara].reset_index(drop=True)

        if operadores_filtrados.empty:
            st.warning(f"No se encontraron operadores para: {busqueda_operador.strip()}")
            st.stop()

        opciones_operador = operadores_filtrados["label"].tolist()
        seleccion_actual = st.session_state.get("nombre_operador_admin")
        if seleccion_actual not in opciones_operador:
            st.session_state["nombre_operador_admin"] = opciones_operador[0]

        selected_operator_label = st.selectbox(
            "Operador a visualizar",
            opciones_operador,
            index=opciones_operador.index(st.session_state["nombre_operador_admin"]),
            key="nombre_operador_admin",
        )
        nombre_operador = str(
            operadores_filtrados.loc[
                operadores_filtrados["label"] == selected_operator_label,
                "operador_codigo",
            ].iloc[0]
        )

    if not nombre_operador.strip():
        st.warning("Debes indicar un operador para procesar la tarea")
        st.stop()

    try:
        tareas = api_get_tasks_by_operator_grouped_by_sku(nombre_operador.strip())
    except Exception as e:
        st.error(f"No se pudo cargar tareas: {e}")
        st.stop()

    tareas = safe_df(tareas)
    inventory_snapshot = get_temporary_inventory_snapshot()
    inventory_loaded = bool(inventory_snapshot.get("loaded"))
    if inventory_loaded and not tareas.empty:
        inventory_df = inventory_snapshot.get("data", pd.DataFrame()).copy()
        tareas = tareas.copy()
        tareas["_sku_inventario"] = tareas["sku"].map(normalize_identifier)
        tareas = tareas.merge(
            inventory_df,
            how="left",
            left_on="_sku_inventario",
            right_on="sku_inventario",
        )
        tareas = tareas.drop(columns=["_sku_inventario", "sku_inventario"], errors="ignore")

    pending_measurements = active_pending_entities("operator_measurement")
    pending_no_stock = active_pending_entities("operator_no_stock")
    if not tareas.empty and pending_measurements:
        tareas = tareas.loc[~tareas["sku"].astype(str).isin(pending_measurements)].reset_index(drop=True)
    if not tareas.empty and pending_no_stock:
        tareas = tareas.copy()
        tareas.loc[tareas["sku"].astype(str).isin(pending_no_stock), "prioridad"] = "sin_stock"

    if not tareas.empty:
        tareas = tareas.copy()
        tareas["_sin_stock_order"] = tareas.apply(row_is_no_stock, axis=1)
        if inventory_loaded:
            tareas["_stock_found"] = tareas["stock_inventario"].notna()
            tareas["_stock_positive"] = tareas["stock_inventario"].fillna(0).astype(float) > 0
            tareas["_stock_recovered"] = tareas["_sin_stock_order"] & tareas["_stock_positive"]

            def inventory_sort_rank(row: pd.Series) -> int:
                if bool(row.get("_stock_recovered", False)):
                    return 0
                if bool(row.get("_stock_positive", False)):
                    return 1
                if not bool(row.get("_stock_found", False)):
                    return 2
                return 3

            tareas["_stock_rank"] = tareas.apply(inventory_sort_rank, axis=1)
            tareas = tareas.sort_values(
                ["_stock_rank", "_sin_stock_order", "sku"],
                ascending=[True, True, True],
            ).drop(columns=["_stock_rank"])
        else:
            tareas = tareas.sort_values(
                ["_sin_stock_order", "sku"],
                ascending=[True, True],
            )

    total_tareas_operador = len(tareas)
    st.metric("Mis SKUs pendientes", total_tareas_operador)

    if inventory_loaded and not tareas.empty:
        con_stock_count = int(tareas["_stock_positive"].sum())
        sin_stock_count = int((tareas["_stock_found"] & ~tareas["_stock_positive"]).sum())
        no_encontrados_count = int((~tareas["_stock_found"]).sum())
        recuperados_count = int(tareas["_stock_recovered"].sum())
        inv_c1, inv_c2, inv_c3, inv_c4 = st.columns(4)
        inv_c1.metric("Con stock", con_stock_count)
        inv_c2.metric("Sin stock", sin_stock_count)
        inv_c3.metric("No encontrados", no_encontrados_count)
        inv_c4.metric("Volvieron a tener stock", recuperados_count)

        inventory_view = st.selectbox(
            "Filtrar según inventario temporal",
            [
                "Todos",
                "Solo con stock",
                "Sin stock",
                "No encontrados",
                "Marcados sin stock que ahora tienen stock",
            ],
            key="operator_inventory_view",
        )
        if inventory_view == "Solo con stock":
            tareas = tareas.loc[tareas["_stock_positive"]].copy()
        elif inventory_view == "Sin stock":
            tareas = tareas.loc[tareas["_stock_found"] & ~tareas["_stock_positive"]].copy()
        elif inventory_view == "No encontrados":
            tareas = tareas.loc[~tareas["_stock_found"]].copy()
        elif inventory_view == "Marcados sin stock que ahora tienen stock":
            tareas = tareas.loc[tareas["_stock_recovered"]].copy()

        st.caption(
            f"Inventario temporal: {inventory_snapshot.get('filename', '')} · "
            f"cargado {inventory_snapshot.get('loaded_at', '')}. "
            "Estos valores no se guardan en Google Sheets."
        )
    elif not inventory_loaded:
        st.info(
            "No hay un inventario temporal cargado. Las tareas se muestran normalmente, "
            "pero sin disponibilidad actual."
        )

    if tareas.empty:
        if total_tareas_operador:
            st.info("No hay SKU que coincidan con el filtro de inventario seleccionado")
        else:
            st.info("No tienes tareas pendientes")
        st.stop()

    def build_operator_label(row: pd.Series) -> str:
        marker = " | MARCADO SIN STOCK" if row_is_no_stock(row) else ""
        stock_marker = ""
        if inventory_loaded:
            stock_marker = f" | STOCK: {format_inventory_stock(row.get('stock_inventario'))}"
        return (
            f"{row['sku']}{marker}{stock_marker} | {row['titulo']} | "
            f"{row.get('publicaciones_count', 0)} publicaciones"
        )

    tareas["label"] = tareas.apply(build_operator_label, axis=1)
    selected_label = st.selectbox("Selecciona SKU", tareas["label"].tolist())
    fila = tareas[tareas["label"] == selected_label].iloc[0]

    if "operador_form_nonce" not in st.session_state:
        st.session_state["operador_form_nonce"] = 0
    form_nonce = st.session_state["operador_form_nonce"]

    st.markdown("### Información actual")
    c1, c2 = st.columns(2)
    with c1:
        st.markdown(f"**SKU:** {fila['sku']}")
        st.markdown(f"**Título:** {fila['titulo']}")
        st.markdown(f"**Publicaciones asociadas:** {fila.get('publicaciones_count', '')}")
        st.markdown(f"**Operador:** {nombre_operador.strip()}")
    with c2:
        st.markdown(f"**Peso ML:** {fila.get('peso_ml_kg', '')} kg")
        st.markdown(
            f"**Dimensiones ML:** {fila.get('alto_ml_cm', '')} x {fila.get('ancho_ml_cm', '')} x {fila.get('profundidad_ml_cm', '')} cm"
        )
        if inventory_loaded:
            st.markdown(f"**Stock temporal:** {format_inventory_stock(fila.get('stock_inventario'))}")
        st.markdown(badge_estado(str(fila.get("estado_actual", ""))), unsafe_allow_html=True)

    sku_sin_stock = row_is_no_stock(fila)
    temporary_stock_value = fila.get("stock_inventario") if inventory_loaded else None
    temporary_stock_found = inventory_loaded and not pd.isna(temporary_stock_value)

    if temporary_stock_found:
        temporary_stock_number = float(temporary_stock_value)
        if temporary_stock_number > 0:
            if sku_sin_stock:
                st.success(
                    f"Este SKU estaba marcado sin stock, pero el informe temporal muestra "
                    f"{format_inventory_stock(temporary_stock_number)} unidades. Se priorizó al inicio de la lista."
                )
            else:
                st.success(
                    f"Stock disponible según informe temporal: "
                    f"{format_inventory_stock(temporary_stock_number)} unidades."
                )
        else:
            st.warning("El informe temporal indica que este SKU continúa sin stock disponible.")
    elif inventory_loaded:
        st.info("Este SKU no aparece en el informe de inventario temporal cargado.")

    if sku_sin_stock and not (temporary_stock_found and float(temporary_stock_value) > 0):
        st.warning(
            "Este SKU está marcado como SIN STOCK. Se mantiene pendiente y queda al final "
            "mientras el inventario temporal no indique disponibilidad."
        )

    with st.form(f"form_sin_stock_{fila['sku']}"):
        observacion_sin_stock = st.text_input(
            "Observación sin stock",
            value="Producto sin stock físico al momento de medir",
            key=f"observacion_sin_stock_{fila['sku']}",
        )
        marcar_sin_stock = st.form_submit_button(
            "Marcar SKU sin stock y enviarlo al final",
            use_container_width=True,
            disabled=sku_sin_stock,
        )

    if marcar_sin_stock:
        try:
            result = api_mark_sku_no_stock(
                sku=str(fila["sku"]),
                operador=nombre_operador.strip(),
                usuario=usuario_actual,
                comentario=observacion_sin_stock,
            )
            remember_pending(
                "operator_no_stock",
                str(fila["sku"]),
                str(result.get("job_id", "")),
            )
            st.success(
                "El SKU se marcó localmente como sin stock y se sincronizará en segundo plano. "
                f"Trabajo: {str(result.get('job_id', ''))[:8]}"
            )
            st.rerun()
        except Exception as e:
            st.error(f"No se pudo marcar sin stock: {e}")

    with st.form("form_medicion_fast"):
        st.markdown("### Ingresar medidas reales")
        col1, col2 = st.columns(2)
        with col1:
            alto = st.number_input(
                "Alto real (cm)",
                min_value=0.1,
                step=0.1,
                format="%.2f",
                key=f"alto_real_fast_{form_nonce}",
            )
            ancho = st.number_input(
                "Ancho real (cm)",
                min_value=0.1,
                step=0.1,
                format="%.2f",
                key=f"ancho_real_fast_{form_nonce}",
            )
        with col2:
            profundidad = st.number_input(
                "Profundidad real (cm)",
                min_value=0.1,
                step=0.1,
                format="%.2f",
                key=f"profundidad_real_fast_{form_nonce}",
            )
            peso = st.number_input(
                "Peso real (kg)",
                min_value=0.001,
                step=0.001,
                format="%.3f",
                key=f"peso_real_fast_{form_nonce}",
            )

        observacion = st.text_area("Observación operador", key=f"observacion_operador_fast_{form_nonce}")
        st.markdown("### Fotos de respaldo")
        foto_alto = st.file_uploader("Foto alto", type=["jpg", "jpeg", "png"], key=f"foto_alto_fast_{form_nonce}")
        foto_ancho = st.file_uploader("Foto ancho", type=["jpg", "jpeg", "png"], key=f"foto_ancho_fast_{form_nonce}")
        foto_profundidad = st.file_uploader("Foto profundidad", type=["jpg", "jpeg", "png"], key=f"foto_profundidad_fast_{form_nonce}")
        foto_peso = st.file_uploader("Foto peso", type=["jpg", "jpeg", "png"], key=f"foto_peso_fast_{form_nonce}")
        submitted = st.form_submit_button("Guardar medición del SKU y subir fotos", use_container_width=True)

    if submitted:
        faltantes = []
        if foto_alto is None:
            faltantes.append("alto")
        if foto_ancho is None:
            faltantes.append("ancho")
        if foto_profundidad is None:
            faltantes.append("profundidad")
        if foto_peso is None:
            faltantes.append("peso")

        if faltantes:
            st.error(f"Faltan fotos obligatorias: {', '.join(faltantes)}")
            st.stop()
        if min(float(alto), float(ancho), float(profundidad), float(peso)) <= 0:
            st.error("Todas las medidas y el peso deben ser mayores que cero")
            st.stop()

        try:
            result = api_save_measurement_with_photos_by_sku(
                sku=str(fila["sku"]),
                operador=nombre_operador.strip(),
                alto_real_cm=float(alto),
                ancho_real_cm=float(ancho),
                profundidad_real_cm=float(profundidad),
                peso_real_kg=float(peso),
                observacion_operador=observacion,
                foto_alto=foto_alto,
                foto_ancho=foto_ancho,
                foto_profundidad=foto_profundidad,
                foto_peso=foto_peso,
            )
            remember_pending(
                "operator_measurement",
                str(fila["sku"]),
                str(result.get("job_id", "")),
            )
            st.session_state["operador_form_nonce"] = st.session_state.get("operador_form_nonce", 0) + 1
            st.success(
                "Medición recibida. Las fotos y los datos se están procesando en segundo plano. "
                f"Trabajo: {str(result.get('job_id', ''))[:8]}"
            )
            st.rerun()
        except Exception as e:
            st.error(f"No se pudo guardar la medición/fotos: {e}")


# =========================================================
# SUPERVISOR
# =========================================================
elif modo == "Supervisor":
    st.title("Módulo Supervisor")

    supervisor_flash = st.session_state.pop("supervisor_flash", None)
    if isinstance(supervisor_flash, dict):
        flash_message = supervisor_flash.get("message", "")
        if supervisor_flash.get("type") == "warning":
            st.warning(flash_message)
        else:
            st.success(flash_message)

    try:
        pendientes = refresh_supervisor_queue(limit=300)
    except Exception as e:
        st.error(f"No se pudo cargar la bandeja: {e}")
        st.stop()

    pendientes = safe_df(pendientes)
    st.metric("SKUs pendientes validación", len(pendientes))

    if pendientes.empty:
        st.info("No hay mediciones pendientes de validación")
        st.stop()

    pendientes["label"] = pendientes.apply(lambda r: f"{r['sku']} | {r['titulo']} | {r.get('publicaciones_count', 0)} publicaciones", axis=1)
    labels = pendientes["label"].tolist()

    selected_label_key = "supervisor_selected_label"
    if st.session_state.get(selected_label_key) not in labels:
        st.session_state[selected_label_key] = labels[0]

    selected_label = st.selectbox("SKU a revisar", labels, key=selected_label_key)
    fila = pendientes[pendientes["label"] == selected_label].iloc[0]

    fallback_case = fila.to_dict()
    detail = {}
    try:
        detail = api_get_case_detail_by_sku(str(fila["sku"]))
    except Exception:
        detail = {}

    case = normalize_case_payload(detail, fallback_case)
    render_case_summary(case)

    comp = pd.DataFrame(
        [
            ["Alto (cm)", case.get("alto_ml_cm", ""), case.get("alto_real_cm", "")],
            ["Ancho (cm)", case.get("ancho_ml_cm", ""), case.get("ancho_real_cm", "")],
            ["Profundidad (cm)", case.get("profundidad_ml_cm", ""), case.get("profundidad_real_cm", "")],
            ["Peso (kg)", case.get("peso_ml_kg", ""), case.get("peso_real_kg", "")],
        ],
        columns=["Campo", "ML", "Real"],
    )
    st.dataframe(comp, use_container_width=True, hide_index=True)

    evid_key = f"show_evid_supervisor_{fila['sku']}"
    st.button(
        "Ver evidencias" if not st.session_state.get(evid_key, False) else "Ocultar evidencias",
        use_container_width=False,
        on_click=toggle_evidencias,
        args=(evid_key,),
        key=f"btn_{evid_key}",
    )

    if st.session_state.get(evid_key, False):
        try:
            evidencias = api_get_evidencias_by_sku(str(fila["sku"]))
        except Exception:
            evidencias = pd.DataFrame(detail.get("evidencias", []))
        render_evidencias(evidencias)

    with st.form(f"supervisor_action_form_{fila['sku']}"):
        comentario = st.text_area("Comentario supervisor", key=f"comentario_supervisor_{fila['sku']}")
        c1, c2 = st.columns(2)
        aprobar = c1.form_submit_button("Aprobar SKU", use_container_width=True)
        devolver = c2.form_submit_button("Solicitar nueva evidencia", use_container_width=True)

    if aprobar:
        try:
            result = api_validate_measurement_by_sku(str(fila["sku"]), usuario_actual, True, comentario)
            remove_supervisor_sku_from_queue(str(fila["sku"]))
            # No modificar st.session_state[selected_label_key] después de crear el selectbox.
            # Streamlit lo bloquea y genera: cannot be modified after the widget is instantiated.
            # En el siguiente rerun, el bloque previo al selectbox corregirá la selección si ya no existe.
            bump_supervisor_queue_version()
            st.session_state["supervisor_flash"] = {
                "type": "success",
                "message": f"Aprobación enviada a segundo plano. Trabajo: {str(result.get('job_id', ''))[:8]}",
            }
        except Exception as e:
            st.error(f"No se pudo aprobar: {e}")
        else:
            st.rerun()

    if devolver:
        try:
            result = api_validate_measurement_by_sku(
                str(fila["sku"]),
                usuario_actual,
                False,
                comentario or "Se solicita nueva evidencia",
            )
            remove_supervisor_sku_from_queue(str(fila["sku"]))
            # No modificar st.session_state[selected_label_key] después de crear el selectbox.
            # Streamlit lo bloquea y genera: cannot be modified after the widget is instantiated.
            # En el siguiente rerun, el bloque previo al selectbox corregirá la selección si ya no existe.
            bump_supervisor_queue_version()
            st.session_state["supervisor_flash"] = {
                "type": "warning",
                "message": f"Solicitud de nueva evidencia enviada a segundo plano. Trabajo: {str(result.get('job_id', ''))[:8]}",
            }
        except Exception as e:
            st.error(f"No se pudo devolver: {e}")
        else:
            st.rerun()


# =========================================================
# ADMINISTRATIVA
# =========================================================
elif modo == "Administrativa":
    st.title("Panel Administrativa")
    bandeja = st.radio("Bandeja", list(BANDEJAS_ADMINISTRATIVA.keys()), horizontal=True)
    estados_bandeja = BANDEJAS_ADMINISTRATIVA[bandeja]

    try:
        cola = api_get_administrative_queue(estados_bandeja, limit=400)
    except Exception as e:
        st.error(f"No se pudo cargar la bandeja: {e}")
        st.stop()

    cola = safe_df(cola)
    pending_admin_keys = active_pending_entities("admin_status")
    if pending_admin_keys and not cola.empty:
        composite = cola.apply(
            lambda row: f"{normalize_identifier(row.get('sku', ''))}::{normalize_identifier(row.get('mlc', ''))}",
            axis=1,
        )
        cola = cola.loc[~composite.isin(pending_admin_keys)].reset_index(drop=True)
    st.metric("Casos en bandeja", len(cola))

    if cola.empty:
        st.info("No hay casos en esta bandeja")
        st.stop()

    with st.form("administrativa_filter_form"):
        texto = st.text_input("Buscar SKU / MLC / título")
        filtro_submit = st.form_submit_button("Aplicar búsqueda", use_container_width=False)

    admina_state = st.session_state.setdefault("administrativa_texto", "")
    if filtro_submit:
        st.session_state["administrativa_texto"] = texto.strip()
        admina_state = texto.strip()
    else:
        admina_state = st.session_state.get("administrativa_texto", "")

    if admina_state:
        mask = (
            cola["sku"].astype(str).str.contains(admina_state, case=False, na=False)
            | cola["mlc"].astype(str).str.contains(admina_state, case=False, na=False)
            | cola["titulo"].astype(str).str.contains(admina_state, case=False, na=False)
        )
        cola = cola[mask].reset_index(drop=True)

    if cola.empty:
        st.info("No se encontraron casos con la búsqueda aplicada")
        st.stop()

    st.subheader("Bandeja de trabajo")
    cols = [
        c for c in [
            "sku", "mlc", "titulo", "estado_actual", "operador_asignado", "supervisor",
            "fecha_validacion", "ticket_ejecutivo"
        ] if c in cola.columns
    ]
    st.dataframe(cola[cols], use_container_width=True, hide_index=True)

    st.markdown("### Cambio masivo de estado")
    bulk_cols = [c for c in cols if c in cola.columns]
    bulk_editor = st.data_editor(
        cola[bulk_cols].assign(seleccionar=False),
        use_container_width=True,
        hide_index=True,
        column_config={"seleccionar": st.column_config.CheckboxColumn("Seleccionar", default=False)},
        disabled=bulk_cols,
        key=f"administrativa_bulk_editor_{bandeja}",
    )
    seleccionados_bulk = bulk_editor[bulk_editor["seleccionar"] == True].copy()  # noqa: E712

    estado_bulk = st.selectbox(
        "Nuevo estado masivo",
        ["listo_para_ejecutivo", "en_gestion_ejecutivo", "resuelto", "rechazado_ml", "rechazado_ejecutivo"],
        key=f"administrativa_bulk_estado_{bandeja}",
    )
    requiere_ticket_bulk = estado_bulk in ["listo_para_ejecutivo", "en_gestion_ejecutivo"]
    ticket_bulk = st.text_input(
        "Ticket ejecutivo masivo",
        key=f"administrativa_bulk_ticket_{bandeja}",
        help="Obligatorio cuando el estado masivo es listo_para_ejecutivo o en_gestion_ejecutivo.",
    )
    comentario_bulk = st.text_area(
        "Comentario masivo",
        height=100,
        key=f"administrativa_bulk_comentario_{bandeja}",
    )

    bulk_state_key = f"administrativa_bulk_validation_{bandeja}"
    c_bulk1, c_bulk2 = st.columns(2)
    with c_bulk1:
        validar_bulk = st.button("Validar selección masiva", use_container_width=True, key=f"btn_validar_bulk_{bandeja}")
    with c_bulk2:
        confirmar_bulk = st.button("Confirmar cambio masivo", use_container_width=True, key=f"btn_confirmar_bulk_{bandeja}")

    if validar_bulk:
        if seleccionados_bulk.empty:
            st.error("Debes seleccionar al menos un caso para el cambio masivo")
            st.session_state.pop(bulk_state_key, None)
        else:
            casos_bulk = seleccionados_bulk.drop(columns=["seleccionar"], errors="ignore").to_dict("records")
            validos_bulk, bloqueados_bulk = validate_bulk_admin_status_change(
                casos=casos_bulk,
                nuevo_estado=estado_bulk,
                comentario=comentario_bulk,
                ticket_ejecutivo=ticket_bulk,
            )
            st.session_state[bulk_state_key] = {
                "nuevo_estado": estado_bulk,
                "ticket": ticket_bulk.strip(),
                "comentario": comentario_bulk.strip(),
                "validos": validos_bulk,
                "bloqueados": bloqueados_bulk,
            }

    bulk_state = st.session_state.get(bulk_state_key)
    if bulk_state:
        st.info(
            f"Validación lista. Válidos: {len(bulk_state.get('validos', []))} | "
            f"Bloqueados: {len(bulk_state.get('bloqueados', []))}"
        )
        if bulk_state.get("bloqueados"):
            bloqueados_df = pd.DataFrame([
                {
                    "sku": str(item.get("caso", {}).get("sku", "")),
                    "mlc": str(item.get("caso", {}).get("mlc", "")),
                    "estado_actual": str(item.get("caso", {}).get("estado_actual", "")),
                    "motivo": str(item.get("motivo", "")),
                }
                for item in bulk_state.get("bloqueados", [])
            ])
            if not bloqueados_df.empty:
                with st.expander("Ver casos bloqueados"):
                    st.dataframe(bloqueados_df, use_container_width=True, hide_index=True)

    if confirmar_bulk:
        bulk_state = st.session_state.get(bulk_state_key)
        if not bulk_state:
            st.error("Primero debes validar la selección masiva")
        elif not bulk_state.get("validos"):
            st.error("No hay casos válidos para actualizar")
        else:
            try:
                validos = bulk_state.get("validos", [])
                result_bulk = api_bulk_update_administrative_status(
                    items=validos,
                    nuevo_estado=str(bulk_state.get("nuevo_estado", "")),
                    usuario=usuario_actual,
                    comentario=str(bulk_state.get("comentario", "")),
                    ticket_ejecutivo=str(bulk_state.get("ticket", "")),
                )
                for caso_bulk in validos:
                    remember_pending(
                        "admin_status",
                        f"{normalize_identifier(caso_bulk.get('sku', ''))}::{normalize_identifier(caso_bulk.get('mlc', ''))}",
                        str(result_bulk.get("job_id", "")),
                    )
                st.session_state.pop(bulk_state_key, None)
                st.success(
                    f"Se enviaron {len(validos)} casos a segundo plano. "
                    f"Trabajo: {str(result_bulk.get('job_id', ''))[:8]}"
                )
                st.rerun()
            except Exception as e:
                st.error(f"No se pudo encolar el cambio masivo: {e}")

    if bandeja == "Pendientes por gestionar":
        st.markdown("### Exportar Excel para ejecutiva")
        seller_id_default = ""
        with st.form("export_ejecutiva_form"):
            seller_id = st.text_input("seller_id", value=str(seller_id_default), key="seller_id_export_fast")
            export_cols = [c for c in ["sku", "mlc", "titulo", "alto_real_cm", "ancho_real_cm", "profundidad_real_cm", "peso_real_kg"] if c in cola.columns]
            export_editor = st.data_editor(
                cola[export_cols].assign(seleccionar=False),
                use_container_width=True,
                hide_index=True,
                column_config={"seleccionar": st.column_config.CheckboxColumn("Seleccionar", default=False)},
                disabled=export_cols,
                key="admin_export_editor_fast",
            )
            preparar_excel = st.form_submit_button("Preparar Excel ejecutiva", use_container_width=True)

        if preparar_excel:
            seleccionados_export = export_editor[export_editor["seleccionar"] == True]  # noqa: E712
            if not seller_id.strip():
                st.error("Debes ingresar seller_id para generar el Excel.")
            elif seleccionados_export.empty:
                st.error("Debes seleccionar al menos un producto.")
            else:
                excel_bytes = build_ejecutiva_excel_bytes(seleccionados_export, seller_id.strip())
                st.download_button(
                    "Descargar Excel ejecutiva",
                    data=excel_bytes,
                    file_name="packaging_para_ejecutiva.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="download_excel_ejecutiva_fast",
                )

    cola["label"] = cola.apply(lambda r: f"{r['sku']} | {r['mlc']} | {r['titulo']}", axis=1)
    selected_label = st.selectbox("Caso", cola["label"].tolist())
    fila = cola[cola["label"] == selected_label].iloc[0]

    fallback_case = fila.to_dict()
    detail = {}
    try:
        detail = api_get_case_detail(str(fila["sku"]), str(fila["mlc"]))
    except Exception:
        detail = {}

    case = normalize_case_payload(detail, fallback_case)
    render_case_summary(case)

    evid_key = f"show_evid_admina_{fila['sku']}_{fila['mlc']}"
    st.button(
        "Ver evidencias" if not st.session_state.get(evid_key, False) else "Ocultar evidencias",
        use_container_width=False,
        on_click=toggle_evidencias,
        args=(evid_key,),
        key=f"btn_{evid_key}",
    )

    if st.session_state.get(evid_key, False):
        try:
            evidencias = api_get_evidencias(str(fila["sku"]), str(fila["mlc"]))
        except Exception:
            evidencias = pd.DataFrame(detail.get("evidencias", []))
        render_evidencias(evidencias)

    st.markdown("### Acción administrativa")
    estado_actual = str(case.get("estado_actual", ""))
    opciones = get_allowed_admin_status_transitions(estado_actual)
    if not opciones:
        st.warning(f"El estado actual no admite cambios administrativos: {estado_actual}")
        st.stop()

    with st.form(f"administrativa_action_form_{fila['sku']}_{fila['mlc']}"):
        nuevo_estado = st.selectbox("Nuevo estado", opciones)
        ticket_default = str(detail.get("case", {}).get("ticket_ejecutivo", "")) if isinstance(detail.get("case"), dict) else str(case.get("ticket_ejecutivo", ""))
        ticket = st.text_input("Ticket ejecutivo", value=ticket_default)
        comentario = st.text_area("Comentario", height=120)
        guardar_gestion = st.form_submit_button("Guardar gestión", use_container_width=True)

    if guardar_gestion:
        error_validacion = validate_admin_status_change(
            estado_actual=estado_actual,
            nuevo_estado=nuevo_estado,
            comentario=comentario,
            ticket_ejecutivo=ticket,
        )
        if error_validacion:
            st.error(error_validacion)
            st.stop()
        try:
            result_status = api_update_status(
                sku=str(fila["sku"]),
                mlc=str(fila["mlc"]),
                nuevo_estado=nuevo_estado,
                usuario=usuario_actual,
                comentario=comentario.strip(),
                ticket_ejecutivo=ticket.strip(),
            )
            remember_pending(
                "admin_status",
                f"{normalize_identifier(fila['sku'])}::{normalize_identifier(fila['mlc'])}",
                str(result_status.get("job_id", "")),
            )
            st.success(
                f"Cambio a {nuevo_estado} enviado a segundo plano. "
                f"Trabajo: {str(result_status.get('job_id', ''))[:8]}"
            )
            st.rerun()
        except Exception as e:
            st.error(f"No se pudo actualizar el caso: {e}")
